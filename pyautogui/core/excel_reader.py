"""Excel 读取：把模板行解析为标准任务数据，供视频/用户流程共用。"""
from __future__ import annotations

import openpyxl

from pyautogui import config


def load_rows(excel_path: str, sheet_name: str, columns: tuple) -> list:
    """返回 [(行号, [各列值]), ...]；跳过目标为空的整行。"""
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        rows = []
        for row in range(config.EXCEL_START_ROW, sheet.max_row + 1):
            values = [sheet.cell(row, col).value for col in columns]
            if values[0] is None or str(values[0]).strip() == "":
                continue
            rows.append((row, values))
        return rows
    finally:
        workbook.close()

